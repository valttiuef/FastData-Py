import os
import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

SRC_DIR = Path(__file__).parent.parent / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from backend.data_db.database import Database
from backend.models import ImportOptions
import frontend.utils.exporting as exporting_module
from frontend.utils.exporting import ExportPlan, _add_excel_chart, execute_export_plan


def test_execute_export_plan_writes_split_csv_files(tmp_path: Path) -> None:
    frame_a = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
    frame_b = pd.DataFrame({"A": [5], "B": [6]})
    destination = tmp_path / "exports.csv"

    plan = ExportPlan(
        kind="dataframes",
        selected_format="csv",
        destination=destination,
        datasets={"First": frame_a, "Second": frame_b},
    )
    outcome = execute_export_plan(plan)

    assert outcome.ok
    assert outcome.open_path == tmp_path
    created = sorted(tmp_path.glob("exports_*_*.csv"))
    assert len(created) == 2

    loaded_first = pd.read_csv(created[0])
    loaded_second = pd.read_csv(created[1])
    assert set(loaded_first.columns) == {"A", "B"}
    assert set(loaded_second.columns) == {"A", "B"}


def test_import_to_export_pipeline_roundtrip_keeps_selected_values(tmp_path: Path) -> None:
    db_path = tmp_path / "pipeline.duckdb"
    csv_path = tmp_path / "input.csv"
    csv_path.write_text(
        "\n".join(
            [
                "Time,Temp,Pressure",
                "2026-01-01 00:00:00,10.0,1.0",
                "2026-01-01 01:00:00,11.0,1.1",
                "2026-01-01 02:00:00,12.0,1.2",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    db = Database(db_path)
    try:
        options = ImportOptions(
            system_name="ExportSystem",
            dataset_name="ExportDataset",
            csv_header_rows=1,
            auto_detect_datetime=False,
            date_column="Time",
        )
        import_ids = db.import_file(csv_path, options)
        assert len(import_ids) == 1

        raw = db.query_raw(system="ExportSystem", dataset="ExportDataset")
        assert isinstance(raw, pd.DataFrame)
        assert not raw.empty

        value_col = "v" if "v" in raw.columns else "value"
        export_frame = raw.loc[:, ["t", value_col]].copy()
        export_frame = export_frame.rename(columns={"t": "Timestamp", value_col: "Value"})
        export_frame["Timestamp"] = pd.to_datetime(export_frame["Timestamp"], errors="coerce")
        export_frame["Value"] = pd.to_numeric(export_frame["Value"], errors="coerce")
        export_frame = export_frame.dropna(subset=["Timestamp", "Value"]).sort_values("Timestamp", kind="stable")
        export_frame = export_frame.reset_index(drop=True)
        assert not export_frame.empty

        destination = tmp_path / "pipeline_export.xlsx"
        plan = ExportPlan(
            kind="dataframes",
            selected_format="excel",
            destination=destination,
            datasets={"Imported data": export_frame},
        )
        outcome = execute_export_plan(plan)
        assert outcome.ok
        assert destination.exists()

        loaded = pd.read_excel(destination, sheet_name="Imported data")
        loaded["Value"] = pd.to_numeric(loaded["Value"], errors="coerce")
        loaded = loaded.dropna(subset=["Value"]).reset_index(drop=True)
        assert len(loaded) == len(export_frame)
        assert float(loaded["Value"].mean()) == pytest.approx(float(export_frame["Value"].mean()))
    finally:
        db.close()


def test_chart_export_unsupported_specs_fall_back_to_data_without_offset(tmp_path: Path) -> None:
    destination = tmp_path / "charts_fallback.xlsx"
    frame = pd.DataFrame(
        {
            "X": [1.0, 2.0],
            "Y": [3.0, 4.0],
            "Z": [5.0, 6.0],
        }
    )
    plan = ExportPlan(
        kind="charts_excel",
        selected_format="excel",
        destination=destination,
        datasets={"Scatter 3D": frame},
        chart_specs={},
        include_charts=True,
        include_data=True,
        chart_first=True,
    )
    outcome = execute_export_plan(plan)

    assert outcome.ok
    loaded = pd.read_excel(destination, sheet_name="Scatter 3D")
    assert list(loaded.columns) == ["X", "Y", "Z"]
    assert len(loaded.index) == 2


def test_add_excel_monthly_chart_sets_non_overlay_title_and_negative_fill_behavior() -> None:
    frame = pd.DataFrame(
        {
            "Period": ["2026-01", "2026-02"],
            "Value": [12.0, -4.0],
        }
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Chart"
    sheet.append(["Period", "Value"])
    sheet.append(["2026-01", 12.0])
    sheet.append(["2026-02", -4.0])

    _add_excel_chart(
        sheet,
        frame,
        {
            "type": "monthly",
            "x_column": "Period",
            "y_columns": ["Value"],
            "title": "Monthly",
        },
        data_sheet=sheet,
        data_start_row=1,
        chart_anchor="A1",
    )

    assert len(sheet._charts) == 1
    chart = sheet._charts[0]
    assert getattr(chart.title, "overlay", None) is False
    assert chart.legend is not None
    assert chart.legend.position == "r"
    assert chart.legend.overlay is False
    assert bool(chart.series)
    assert chart.series[0].invertIfNegative is False


def test_add_excel_time_series_chart_sets_non_overlay_right_legend() -> None:
    frame = pd.DataFrame(
        {
            "t": pd.to_datetime(["2026-01-01", "2026-01-02"]),
            "A": [1.0, 2.0],
            "B": [3.0, 4.0],
        }
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Series"
    sheet.append(["t", "A", "B"])
    sheet.append([pd.Timestamp("2026-01-01"), 1.0, 3.0])
    sheet.append([pd.Timestamp("2026-01-02"), 2.0, 4.0])

    _add_excel_chart(
        sheet,
        frame,
        {
            "type": "time_series",
            "x_column": "t",
            "y_columns": ["A", "B"],
            "title": "Series",
        },
        data_sheet=sheet,
        data_start_row=1,
        chart_anchor="A1",
    )

    assert len(sheet._charts) == 1
    chart = sheet._charts[0]
    assert getattr(chart.title, "overlay", None) is False
    assert chart.legend is not None
    assert chart.legend.position == "r"
    assert chart.legend.overlay is False


def test_add_excel_scatter_chart_hides_legend() -> None:
    frame = pd.DataFrame(
        {
            "X": [1.0, 2.0, 3.0],
            "Y": [2.0, 3.0, 4.0],
        }
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Scatter"
    sheet.append(["X", "Y"])
    sheet.append([1.0, 2.0])
    sheet.append([2.0, 3.0])
    sheet.append([3.0, 4.0])

    _add_excel_chart(
        sheet,
        frame,
        {
            "type": "scatter",
            "x_column": "X",
            "y_columns": ["Y"],
            "title": "Scatter",
        },
        data_sheet=sheet,
        data_start_row=1,
        chart_anchor="A1",
    )

    assert len(sheet._charts) == 1
    chart = sheet._charts[0]
    assert chart.legend is None


def test_chart_export_supported_specs_keep_chart_first_data_offset(tmp_path: Path) -> None:
    destination = tmp_path / "charts_supported.xlsx"
    frame = pd.DataFrame(
        {
            "Period": ["2026-01", "2026-02"],
            "Value": [1.0, 2.0],
        }
    )
    plan = ExportPlan(
        kind="charts_excel",
        selected_format="excel",
        destination=destination,
        datasets={"Monthly": frame},
        chart_specs={
            "Monthly": {
                "type": "monthly",
                "x_column": "Period",
                "y_columns": ["Value"],
                "title": "Monthly",
            }
        },
        include_charts=True,
        include_data=True,
        chart_first=True,
    )
    outcome = execute_export_plan(plan)

    assert outcome.ok
    workbook = load_workbook(destination)
    sheet = workbook["Monthly"]
    assert sheet["A1"].value is None
    assert sheet["A20"].value == "Period"
    assert len(sheet._charts) == 1


def test_execute_export_plan_reports_file_in_use_cleanly(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    destination = tmp_path / "locked.xlsx"
    frame = pd.DataFrame({"A": [1.0]})
    plan = ExportPlan(
        kind="dataframes",
        selected_format="excel",
        destination=destination,
        datasets={"Data": frame},
    )

    monkeypatch.setattr(exporting_module, "_is_destination_writable", lambda _path: False)
    outcome = execute_export_plan(plan)

    assert not outcome.ok
    assert "currently in use" in str(outcome.message).lower()


def test_excel_export_autosizes_columns_for_long_headers(tmp_path: Path) -> None:
    destination = tmp_path / "autosize.xlsx"
    long_header = "Temperature sensor very long feature name"
    frame = pd.DataFrame({long_header: [1.0, 2.0], "Short": [3.0, 4.0]})
    plan = ExportPlan(
        kind="dataframes",
        selected_format="excel",
        destination=destination,
        datasets={"Data": frame},
    )
    outcome = execute_export_plan(plan)

    assert outcome.ok
    workbook = load_workbook(destination)
    sheet = workbook["Data"]
    assert float(sheet.column_dimensions["A"].width or 0.0) >= 20.0
