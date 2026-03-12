
from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
import re

import pandas as pd
from PySide6.QtWidgets import QFileDialog, QWidget
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.utils import get_column_letter

from ..localization import tr
from .file_dialog_history import get_dialog_directory, remember_dialog_path
import logging

logger = logging.getLogger(__name__)


def _slugify(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", (value or "").strip())
    return text.strip("_") or "export"


def _sheet_name(value: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]+", "_", value)
    cleaned = cleaned.strip() or "Sheet"
    return cleaned[:31]


@dataclass(frozen=True)
class ExportPlan:
    kind: str
    selected_format: str
    destination: Path
    datasets: dict[str, pd.DataFrame]
    chart_specs: dict[str, dict[str, object]] | None = None
    include_charts: bool = True
    include_data: bool = True
    chart_first: bool = True


@dataclass(frozen=True)
class ExportOutcome:
    ok: bool
    message: str
    open_path: Path | None = None


def _file_in_use_message(path: Path) -> str:
    return tr("File is currently in use: {path}. Close it and try again.").format(path=str(path))


def _is_destination_writable(path: Path) -> bool:
    if path.exists() and path.is_dir():
        return False
    if not path.exists():
        return True
    try:
        with path.open("a+b"):
            pass
        return True
    except PermissionError:
        return False
    except OSError:
        return True


def _chart_spec_can_render(frame: pd.DataFrame, spec: dict[str, object] | None) -> bool:
    if frame is None or frame.empty:
        return False
    columns = [str(c) for c in frame.columns]
    if len(columns) <= 1:
        return False

    spec = spec or {}
    chart_type = str(spec.get("type") or "").lower()
    if chart_type not in {"monthly", "time_series", "scatter"}:
        return False

    x_column = str(spec.get("x_column") or "")
    y_columns = spec.get("y_columns")
    if not x_column or x_column not in columns:
        x_column = columns[0]
    if not y_columns or not isinstance(y_columns, (list, tuple)):
        y_columns = [c for c in columns if c != x_column]

    y_col_idxs = [columns.index(col) + 1 for col in y_columns if col in columns and col != x_column]
    if not y_col_idxs:
        return False
    if chart_type == "scatter":
        return len(y_col_idxs) == 1
    return True


def _apply_chart_title_layout(chart) -> None:
    try:
        title_obj = getattr(chart, "title", None)
        if title_obj is not None:
            title_obj.overlay = False
    except Exception:
        logger.warning("Exception in _apply_chart_title_layout", exc_info=True)


def _apply_chart_legend_layout(chart) -> None:
    try:
        legend = getattr(chart, "legend", None)
        if legend is None:
            return
        legend.position = "r"
        legend.overlay = False
    except Exception:
        logger.warning("Exception in _apply_chart_legend_layout", exc_info=True)


def _autosize_worksheet_columns(
    sheet,
    *,
    min_width: int = 12,
    max_width: int = 56,
    max_rows: int = 200,
) -> None:
    if sheet is None:
        return
    try:
        row_count = int(getattr(sheet, "max_row", 0) or 0)
        col_count = int(getattr(sheet, "max_column", 0) or 0)
    except Exception:
        return
    if row_count <= 0 or col_count <= 0:
        return

    scan_rows = max(1, min(row_count, int(max_rows)))
    for col_idx in range(1, col_count + 1):
        max_len = 0
        for row_idx in range(1, scan_rows + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            text = str(value)
            if len(text) > max_len:
                max_len = len(text)
        width = max(min_width, min(max_width, max_len + 2))
        try:
            sheet.column_dimensions[get_column_letter(col_idx)].width = float(width)
        except Exception:
            logger.warning("Exception in _autosize_worksheet_columns", exc_info=True)


# @ai(gpt-5, codex, refactor, 2026-03-11)
def prepare_dataframes_export_plan(
    *,
    parent: QWidget | None,
    title: str,
    selected_format: str,
    datasets: dict[str, pd.DataFrame],
) -> ExportPlan | None:
    if not datasets:
        return None
    clean = {name: (df if isinstance(df, pd.DataFrame) else pd.DataFrame(df)) for name, df in datasets.items()}
    mode = str(selected_format or "csv").lower()
    if mode == "excel":
        suggested = get_dialog_directory(parent, "export", Path.cwd()) / f"{_slugify(title)}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            parent,
            title,
            str(suggested),
            tr("Excel files (*.xlsx);;All files (*.*)"),
        )
        if not path:
            return None
        dest = Path(path)
        if dest.suffix.lower() != ".xlsx":
            dest = dest.with_suffix(".xlsx")
        remember_dialog_path(parent, "export", dest)
        return ExportPlan(kind="dataframes", selected_format=mode, destination=dest, datasets=clean)

    suggested = get_dialog_directory(parent, "export", Path.cwd()) / f"{_slugify(title)}.csv"
    path, _ = QFileDialog.getSaveFileName(
        parent,
        title,
        str(suggested),
        tr("CSV files (*.csv);;All files (*.*)"),
    )
    if not path:
        return None
    dest = Path(path)
    if dest.suffix.lower() != ".csv":
        dest = dest.with_suffix(".csv")
    remember_dialog_path(parent, "export", dest)
    return ExportPlan(kind="dataframes", selected_format=mode, destination=dest, datasets=clean)


# @ai(gpt-5, codex, refactor, 2026-03-11)
def prepare_charts_excel_export_plan(
    *,
    parent: QWidget | None,
    title: str,
    datasets: dict[str, pd.DataFrame],
    chart_specs: dict[str, dict[str, object]] | None = None,
    include_charts: bool = True,
    include_data: bool = True,
    chart_first: bool = True,
) -> ExportPlan | None:
    if not datasets:
        return None
    suggested = get_dialog_directory(parent, "export", Path.cwd()) / f"{_slugify(title)}.xlsx"
    path, _ = QFileDialog.getSaveFileName(
        parent,
        title,
        str(suggested),
        tr("Excel files (*.xlsx);;All files (*.*)"),
    )
    if not path:
        return None
    dest = Path(path)
    if dest.suffix.lower() != ".xlsx":
        dest = dest.with_suffix(".xlsx")
    remember_dialog_path(parent, "export", dest)
    clean = {name: (df if isinstance(df, pd.DataFrame) else pd.DataFrame(df)) for name, df in datasets.items()}
    return ExportPlan(
        kind="charts_excel",
        selected_format="excel",
        destination=dest,
        datasets=clean,
        chart_specs=chart_specs or {},
        include_charts=bool(include_charts),
        include_data=bool(include_data),
        chart_first=bool(chart_first),
    )


# @ai(gpt-5, codex, refactor, 2026-03-11)
def execute_export_plan(plan: ExportPlan) -> ExportOutcome:
    clean = {name: (df if isinstance(df, pd.DataFrame) else pd.DataFrame(df)) for name, df in plan.datasets.items()}
    if not clean:
        return ExportOutcome(ok=False, message=tr("Nothing selected to export."))

    mode = str(plan.selected_format or "csv").lower()
    if plan.kind == "charts_excel":
        return _execute_chart_excel_export(clean=clean, plan=plan)
    return _execute_dataframes_export(clean=clean, plan=plan, mode=mode)


# @ai(gpt-5, codex, refactor, 2026-03-12)
def _execute_dataframes_export(*, clean: dict[str, pd.DataFrame], plan: ExportPlan, mode: str) -> ExportOutcome:
    if mode == "excel":
        dest = plan.destination
        if not _is_destination_writable(dest):
            return ExportOutcome(ok=False, message=_file_in_use_message(dest))
        try:
            with pd.ExcelWriter(dest, engine="openpyxl") as writer:
                for name, frame in clean.items():
                    sheet_name = _sheet_name(name)
                    frame.to_excel(writer, sheet_name=sheet_name, index=False)
                    _autosize_worksheet_columns(writer.sheets.get(sheet_name))
        except PermissionError:
            return ExportOutcome(ok=False, message=_file_in_use_message(dest))
        return ExportOutcome(
            ok=True,
            message=tr("Exported {count} dataset(s) to {path}.").format(count=len(clean), path=str(dest)),
            open_path=dest,
        )

    base = plan.destination
    if len(clean) == 1:
        if not _is_destination_writable(base):
            return ExportOutcome(ok=False, message=_file_in_use_message(base))
        try:
            next(iter(clean.values())).to_csv(base, index=False)
        except PermissionError:
            return ExportOutcome(ok=False, message=_file_in_use_message(base))
        return ExportOutcome(
            ok=True,
            message=tr("Exported data to {path}.").format(path=str(base)),
            open_path=base,
        )

    written = []
    stem = base.stem
    parent_dir = base.parent
    for idx, (name, frame) in enumerate(clean.items(), start=1):
        file_name = f"{stem}_{idx:02d}_{_slugify(name)}.csv"
        dest = parent_dir / file_name
        if not _is_destination_writable(dest):
            return ExportOutcome(ok=False, message=_file_in_use_message(dest))
        try:
            frame.to_csv(dest, index=False)
        except PermissionError:
            return ExportOutcome(ok=False, message=_file_in_use_message(dest))
        written.append(str(dest))

    return ExportOutcome(
        ok=True,
        message=tr("Exported {count} CSV files to {folder}.").format(
            count=len(written), folder=str(parent_dir)
        ),
        open_path=parent_dir,
    )


# @ai(gpt-5, codex, refactor, 2026-03-12)
def _execute_chart_excel_export(*, clean: dict[str, pd.DataFrame], plan: ExportPlan) -> ExportOutcome:
    dest = plan.destination
    if not _is_destination_writable(dest):
        return ExportOutcome(ok=False, message=_file_in_use_message(dest))
    specs = plan.chart_specs or {}
    include_charts = bool(plan.include_charts)
    include_data = bool(plan.include_data)
    chart_first = bool(plan.chart_first)
    unsupported_charts: list[str] = []

    try:
        with pd.ExcelWriter(dest, engine="openpyxl") as writer:
            for name, frame in clean.items():
                sheet_name = _sheet_name(name)
                spec = specs.get(name) or {}
                can_render_chart = include_charts and _chart_spec_can_render(frame, spec)
                if include_charts and not can_render_chart:
                    unsupported_charts.append(sheet_name)

                if include_charts and not include_data and can_render_chart:
                    data_sheet_name = _sheet_name(f"{name} Data")
                    frame.to_excel(writer, sheet_name=data_sheet_name, index=False)
                    data_sheet = writer.sheets.get(data_sheet_name)
                    _autosize_worksheet_columns(data_sheet)
                    chart_sheet = writer.book.create_sheet(sheet_name)
                    if frame is not None and not frame.empty:
                        _add_excel_chart(
                            chart_sheet,
                            frame,
                            spec,
                            data_sheet=data_sheet,
                            data_start_row=1,
                            chart_anchor="A1",
                        )
                    try:
                        data_sheet.sheet_state = "hidden"
                    except Exception:
                        logger.warning("Exception in _execute_chart_excel_export", exc_info=True)
                    continue

                data_start_row = 1
                chart_anchor = "A1"
                include_data_for_sheet = include_data or not can_render_chart
                if can_render_chart and include_data_for_sheet and chart_first:
                    data_start_row = 20

                if include_data_for_sheet:
                    frame.to_excel(writer, sheet_name=sheet_name, index=False, startrow=data_start_row - 1)
                    data_sheet = writer.sheets.get(sheet_name)
                    _autosize_worksheet_columns(data_sheet)
                else:
                    data_sheet = writer.book.create_sheet(sheet_name)

                if can_render_chart and frame is not None and not frame.empty:
                    _add_excel_chart(
                        data_sheet,
                        frame,
                        spec,
                        data_sheet=data_sheet,
                        data_start_row=data_start_row,
                        chart_anchor=chart_anchor,
                    )
    except PermissionError:
        return ExportOutcome(ok=False, message=_file_in_use_message(dest))

    message = tr("Exported {count} dataset(s) to {path}.").format(count=len(clean), path=str(dest))
    if unsupported_charts:
        message += " " + tr("Unsupported chart types were exported as data only.")
    return ExportOutcome(
        ok=True,
        message=message,
        open_path=dest,
    )


def export_dataframes(
    *,
    parent: QWidget | None,
    title: str,
    selected_format: str,
    datasets: dict[str, pd.DataFrame],
) -> tuple[bool, str]:
    if not datasets:
        return False, tr("Nothing selected to export.")
    plan = prepare_dataframes_export_plan(
        parent=parent,
        title=title,
        selected_format=selected_format,
        datasets=datasets,
    )
    if plan is None:
        return False, ""
    outcome = execute_export_plan(plan)
    return outcome.ok, outcome.message


def export_charts_excel(
    *,
    parent: QWidget | None,
    title: str,
    datasets: dict[str, pd.DataFrame],
    chart_specs: dict[str, dict[str, object]] | None = None,
    include_charts: bool = True,
    include_data: bool = True,
    chart_first: bool = True,
) -> tuple[bool, str]:
    if not datasets:
        return False, tr("Nothing selected to export.")
    plan = prepare_charts_excel_export_plan(
        parent=parent,
        title=title,
        datasets=datasets,
        chart_specs=chart_specs,
        include_charts=include_charts,
        include_data=include_data,
        chart_first=chart_first,
    )
    if plan is None:
        return False, ""
    outcome = execute_export_plan(plan)
    return outcome.ok, outcome.message


# @ai(gpt-5, codex, refactor, 2026-03-12)
def _add_excel_chart(
    sheet,
    frame: pd.DataFrame,
    spec: dict[str, object],
    *,
    data_sheet=None,
    data_start_row: int = 1,
    chart_anchor: str = "A1",
) -> None:
    if sheet is None:
        return
    chart_type = str(spec.get("type") or "").lower()
    if chart_type not in {"monthly", "time_series", "scatter"}:
        return

    columns = [str(c) for c in frame.columns]
    if not columns:
        return

    row_count = len(frame.index)
    col_count = len(columns)
    if row_count <= 0 or col_count <= 1:
        return

    data_start_row = max(1, int(data_start_row))
    data_end_row = data_start_row + row_count

    if data_sheet is None:
        data_sheet = sheet

    x_column = str(spec.get("x_column") or "")
    y_columns = spec.get("y_columns")
    if not x_column or x_column not in columns:
        x_column = columns[0]
    if not y_columns or not isinstance(y_columns, (list, tuple)):
        y_columns = [c for c in columns if c != x_column]

    if not y_columns:
        return

    x_col_idx = columns.index(x_column) + 1
    y_col_idxs = [columns.index(col) + 1 for col in y_columns if col in columns and col != x_column]
    if not y_col_idxs:
        return

    if chart_type == "scatter":
        if len(y_col_idxs) != 1:
            return
        chart = ScatterChart()
        chart.title = str(spec.get("title") or "Scatter")
        chart.legend = None
        xvalues = Reference(data_sheet, min_col=x_col_idx, min_row=data_start_row + 1, max_row=data_end_row)
        yvalues = Reference(data_sheet, min_col=y_col_idxs[0], min_row=data_start_row + 1, max_row=data_end_row)
        series = Series(
            yvalues,
            xvalues,
            title=data_sheet.cell(row=data_start_row, column=y_col_idxs[0]).value,
        )
        try:
            series.marker = Marker(symbol="circle")
            series.marker.size = 5
            series.graphicalProperties.line.noFill = True
        except Exception:
            logger.warning("Exception in _add_excel_chart", exc_info=True)
        chart.series.append(series)
        _apply_chart_title_layout(chart)
    else:
        if chart_type == "monthly":
            chart = BarChart()
        else:
            chart = LineChart()
        chart.title = str(spec.get("title") or "Chart")
        data_ref = Reference(
            data_sheet,
            min_col=min(y_col_idxs),
            max_col=max(y_col_idxs),
            min_row=data_start_row,
            max_row=data_end_row,
        )
        chart.add_data(data_ref, titles_from_data=True)
        categories = Reference(
            data_sheet,
            min_col=x_col_idx,
            min_row=data_start_row + 1,
            max_row=data_end_row,
        )
        chart.set_categories(categories)
        _apply_chart_title_layout(chart)
        _apply_chart_legend_layout(chart)
        if chart_type == "monthly":
            try:
                for series in chart.series:
                    series.invertIfNegative = False
            except Exception:
                logger.warning("Exception in _add_excel_chart", exc_info=True)

    sheet.add_chart(chart, chart_anchor)


__all__ = [
    "ExportOutcome",
    "ExportPlan",
    "execute_export_plan",
    "export_dataframes",
    "export_charts_excel",
    "prepare_charts_excel_export_plan",
    "prepare_dataframes_export_plan",
]
